#
# Analysis the Expedition routing file and grade wind and angles
#

import os
import sys
from pathlib import Path
import datetime as dt
from datetime import timedelta
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
#from openpyxl.cell import get_column_letter
from openpyxl.utils import get_column_letter
#from openpyxl.styles import Font
from openpyxl.styles import Color, PatternFill, Font, Border

# columns for TWS and TWA
TWSCOL = 4
TWACOL = 6

# The input is the CSV file - default for testing
fname = "fastnet 210327 leg1"

# and the output an Excel spreashsheet
exWorkbook = Workbook() # for Expedition
#exSheet = exWorkbook.active
exSheet = exWorkbook.worksheets[0]
exSheet.title = "Route Summary"

#print(f"Arguments count: {len(sys.argv)}")
for i, arg in enumerate(sys.argv):
    #print(f"Argument {i:>6}: {arg}")
    if (i == 1):
        fname = arg

ffile = Path(fname)
if ffile.is_file():
    print(f"Processing file {fname}")
else:
    fname = fname + '.csv'
    ffile = Path(fname)
    if ffile.is_file():
        print(f"Processing file {fname}")
    else:
        print(f"File {fname} not found")
        exit(1)


# opens the csv file defined below.
#inputfile = open("fastnet 210327 leg1.csv", 'rt')
inputfile = open(fname, 'rt')

# creates the reader object
reader = csv.reader(inputfile)
rownum = 0

# define the colours for shading
# see https://www.rapidtables.com/web/color/RGB_Color.html
Bft1Fill = PatternFill(start_color='778899', end_color='778899', fill_type='solid')
Bft2Fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
Bft3Fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
Bft4Fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
Bft5Fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
Bft6Fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
Bft7Fill = PatternFill(start_color='FA8072', end_color='FA8072', fill_type='solid')
Bft8Fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
Bft9Fill = PatternFill(start_color='DB7093', end_color='DB7093', fill_type='solid')
Bft10Fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')

colours = Bft1Fill, Bft2Fill, Bft3Fill, Bft3Fill, Bft5Fill, Bft6Fill, Bft7Fill, Bft8Fill, Bft9Fill
# add - list of dicts
Winds = [
    {'low': 0.0, 'high': 4.0, 'fill': Bft1Fill}, # 1
    {'low': 4.0, 'high': 7.0, 'fill': Bft2Fill}, # 2
    {'low': 7.0, 'high': 11.0, 'fill': Bft3Fill}, # 3
    {'low': 11.0, 'high': 17.0, 'fill': Bft4Fill},
    {'low': 17.0, 'high': 22.0, 'fill': Bft5Fill},
    {'low': 22.0, 'high': 28.0, 'fill': Bft6Fill},
    {'low': 28.0, 'high': 34.0, 'fill': Bft7Fill},
    {'low': 34.0, 'high': 41.0, 'fill': Bft8Fill},
    {'low': 41.0, 'high': 48.0, 'fill': Bft9Fill}
]

dateDelta = timedelta(days=0, hours=0, minutes=0, seconds=0)
Angles = [
    {"low": 0, "high": 30, "tot": 0, "percent": 0}, # high
    {"low": 30, "high": 60, "tot": 0, "percent": 0}, # close hauled
    {"low": 60, "high": 90, "tot": 0, "percent": 0}, # cracked
    {"low": 90, "high": 120, "tot": 0, "percent": 0}, # reaching
    {"low": 120, "high": 150, "tot": 0, "percent": 0}, 
    {"low": 150, "high": 180, "tot": 0, "percent": 0} # running
]

twsTotal = 0.0
twsCnt = 0
offset = 0
bucket = [0, 0, 0]
# add rows to spreadsheet from csv file
# csv starts from 0,0 and worksheet 1,1 so '+1' needed
prevDttm = 0
dttmInv = 0
totalDttm = 0
#dateDelta = timedelta(days=0, hours=0, minutes=0, seconds=0)
for row_index, row in enumerate(reader):
    col = 0 # used to count output columns
    for column_index, cell in enumerate(row):
        if (column_index >= 15 and column_index <= 28): # skip Rain/MSLP
            continue

        if (row_index > 0 and col == 0): # date such as 29-Mar-21 03:00
            if (cell.find(":") > 0):
                dttm = dt.datetime.strptime(cell, "%d-%b-%y %H:%M")
                offset = dttm.timestamp() # convert to epoc
                #print (dttm)
                if (prevDttm != 0): # there is a previous
                    dttmInv = offset - prevDttm

                prevDttm = offset
                totalDttm = totalDttm + offset
            #SheetName.cell(row=i,column=1).value = dttm

        col = col + 1
        column_letter = get_column_letter((column_index + 1))

        if ('(' in cell): # deals with '()' around numbers
            cell = cell[1:len(cell) - 1]

        # this is the complicated bit - convert valid field to ints/floats
        if (cell.isnumeric()): # doesn't cover '.'
            cell = int(cell)
        elif (len(cell) > 0 and cell[0] == '-'): # neg numbers
            cell = int(cell)
        elif (len(cell) > 0 and cell.find(' ') == -1 and
                                 cell.find('nm') == -1 and cell.find('.') > 0):
            cell = float(cell)

        exSheet.cell((row_index + 1), (col)).value = cell
        #exSheet.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

        # if TWS and not top row and has value
                #isinstance(cell, str) and len(cell) > 0): # then TWS
        if (column_index == TWSCOL and row_index != 0 and isinstance (cell, float) and cell > 0.0):
            twsTotal = twsTotal + float(cell)
            twsCnt = twsCnt + 1

            # shade TWS based on wind strenght
            for bft in Winds:
                if (cell >= bft["low"] and cell < bft["high"]):
                    exSheet.cell((row_index + 1), (column_index + 1)).fill = bft["fill"]
                    continue;

        # note the time at wind angle
        if (column_index == TWACOL and row_index !=0 and (isinstance (cell, float) or isinstance (cell, int)) ):
            ang = float (cell) # redundant
            if (ang < 0): ang = ang + 180 # make positive
            # angles are 0 to 180 and 0 to -180
            for rng in Angles:
                if (ang > rng["low"] and ang <= rng["high"]):
                    #print (dttmInv)
                    rng["percent"] = rng["percent"] + offset # how long at this angle
                    continue;
# end of rows/cols

# added the avg wind at the bollot
exSheet.cell((twsCnt + 2), (TWSCOL)).value = 'Average:'
exSheet.cell((twsCnt + 2), (TWSCOL + 1)).value = twsTotal/twsCnt
exSheet.cell((twsCnt + 2), (TWSCOL + 1)).number_format = '0.0'
exSheet.cell((twsCnt + 2), (TWSCOL + 1)).font = Font(bold = True)
# format 1 decimal place - ws['A1'].number_format = '0.00%'

# And the wind buckets - 
clr = 0
for rng in Angles:
    per = rng["percent"]/totalDttm
    #per = tot/totalDttm 
    exSheet.cell(twsCnt + 6 + clr, 6).value = clr + 1
    exSheet.cell(twsCnt + 6 + clr, 7).value = str(int(rng["low"])+1) + "-" + str(int(rng["high"]))
    exSheet.cell(twsCnt + 6 + clr, 8).value = per # rng["percent"]
    exSheet.cell(twsCnt + 6 + clr, 8).number_format = '00.0%'
    clr = clr + 1

# added the wind scaling @53 x 12 - the scaling
exSheet.cell(twsCnt + 6, 11).value = "Wind (Bft)"

clr = 0
for clr in range(9):
    exSheet.cell(twsCnt + 6 + clr, 12).value = clr + 1
    exSheet.cell(twsCnt + 6 + clr, 13).value = str(int(Winds[clr]["low"])) + "-" + str(int(Winds[clr]["high"])-1)
    exSheet.cell(twsCnt + 6 + clr, 13).fill = Winds[clr]["fill"]


print ("TWS ", twsTotal, twsTotal/twsCnt, twsCnt) # Avg wind
oname, fext = os.path.splitext(fname)
exWorkbook.save(oname + ".xlsx")

# end of file

