#
# Analysis the Expedition routing file and grade wind and angles
#
# this reads the filtered output file and summarises
# file should be saved with TWS steps of 1 and TWA steps of 10
#

import os
import sys
from pathlib import Path
import datetime as dt
from datetime import timedelta
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
#from openpyxl.cell import get_column_letter
from openpyxl.utils import get_column_letter
#from openpyxl.styles import Font
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Alignment
# added for conditional shading
from openpyxl.styles import colors
from openpyxl.formatting.rule import ColorScaleRule

# columns for TWS and TWA
TWSCOL = 4
TWACOL = 6

# The input is the CSV file - default for testing
fname = "fastnet 210328 filt"

#print(f"Arguments count: {len(sys.argv)}")
for i, arg in enumerate(sys.argv):
    #print(f"Argument {i:>6}: {arg}")
    if (i == 1):
        fname = arg

ffile = Path(fname)
if ffile.is_file():
    print(f"Processing file {fname}")
else:
    fname = fname + '.csv'
    ffile = Path(fname)
    if ffile.is_file():
        print(f"Processing file {fname}")
    else:
        print(f"File {fname} not found")
        exit(1)

# and the output an Excel spreashsheet & name the sheet
exWorkbook = Workbook() # for Expedition
#exSheet = exWorkbook.active
exSheet = exWorkbook.worksheets[0]
exSheet.title = "Filter Summary" # different to default

# opens the csv file defined below.
inputfile = open(fname, 'rt')

# creates the reader object
reader = csv.reader(inputfile)
rownum = 0

#import winds.py # the shading file
# need to externalised these
# define the colours for shading
# see https://www.rapidtables.com/web/color/RGB_Color.html
# and colours taken from windy - https://community.windy.com/topic/8748/wind-speed-values-on-map/3
Bft1Fill = PatternFill(start_color='778899', end_color='778899', fill_type='solid')
Bft2Fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
Bft3Fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
Bft4Fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
Bft5Fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
Bft6Fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
Bft7Fill = PatternFill(start_color='FA8072', end_color='FA8072', fill_type='solid')
Bft8Fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
Bft9Fill = PatternFill(start_color='DB7093', end_color='DB7093', fill_type='solid')
Bft10Fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')

colours = Bft1Fill, Bft2Fill, Bft3Fill, Bft4Fill, Bft5Fill, Bft6Fill, Bft7Fill, Bft8Fill, Bft9Fill, Bft10Fill
# add - list of dicts
Winds = [
    {'low': 0.0, 'high': 4.0, 'fill': Bft1Fill}, # 1
    {'low': 4.0, 'high': 7.0, 'fill': Bft2Fill}, # 2
    {'low': 7.0, 'high': 11.0, 'fill': Bft3Fill}, # 3
    {'low': 11.0, 'high': 17.0, 'fill': Bft4Fill},
    {'low': 17.0, 'high': 22.0, 'fill': Bft5Fill},
    {'low': 22.0, 'high': 28.0, 'fill': Bft6Fill},
    {'low': 28.0, 'high': 34.0, 'fill': Bft7Fill},
    {'low': 34.0, 'high': 41.0, 'fill': Bft8Fill},
    {'low': 41.0, 'high': 48.0, 'fill': Bft9Fill}
]

# Hours is a copy of BFT for now - tweak later
# Excel conditional shading defaults to 3 coolurs red, yellow and green
Hours = [
    {'low': 0.0, 'high': 0.5, 'fill': Bft1Fill}, # 1
    {'low': 0.5, 'high': 1.0, 'fill': Bft2Fill}, # 2
    {'low': 1.0, 'high': 2.0, 'fill': Bft3Fill}, # 3
    {'low': 2.0, 'high': 4.0, 'fill': Bft4Fill},
    {'low': 4.0, 'high': 8.0, 'fill': Bft5Fill},
    {'low': 8.0, 'high': 16.0, 'fill': Bft6Fill},
    {'low': 16.0, 'high': 32.0, 'fill': Bft7Fill},
    {'low': 32.0, 'high': 64.0, 'fill': Bft8Fill},
    {'low': 64.0, 'high': 128.0, 'fill': Bft9Fill}
]

dateDelta = timedelta(days=0, hours=0, minutes=0, seconds=0)
Angles = [
    {"low": 0, "high": 30, "tot": 0, "percent": 0}, # high
    {"low": 30, "high": 60, "tot": 0, "percent": 0}, # close hauled
    {"low": 60, "high": 90, "tot": 0, "percent": 0}, # cracked
    {"low": 90, "high": 120, "tot": 0, "percent": 0}, # reaching
    {"low": 120, "high": 150, "tot": 0, "percent": 0}, 
    {"low": 150, "high": 180, "tot": 0, "percent": 0} # running
]

totHours = 0.0
twsTotal = 0.0
twsCnt = 0
offset = 0
bucket = [0, 0, 0]
# add rows to spreadsheet from csv file
# csv starts from 0,0 and worksheet 1,1 so '+1' needed
prevDttm = 0
dttmInv = 0
totalDttm = 0

# construct the 2 dim array for the times of 40 x 8
rows = 40
cols = 18
Times = [[0.0] * cols for i in range(rows)] 

# shade based on hours
def shadeHrs(val):
    for hr in Hours:
        if (val >= hr["low"] and val < hr["high"]):
            return (hr["fill"]);
            #continue;
    #return 'black'
# end shadeHrs

# function for shading wind
def shadeWind(wind):
    for bft in Winds:
        if (wind >= bft["low"] and wind < bft["high"]):
            return (bft["fill"]);
            #continue;
    #return 'black'
# end shadeWind

# the boundaries for the tws/twas area of the csv file
firstRow = 8
lastRow = 49
lastCol = 18

#dateDelta = timedelta(days=0, hours=0, minutes=0, seconds=0)
# iterate over all the rows in the csv file
for row_index, row in enumerate(reader):
    rowHrs = 0
    offset = 0;
    col = 0 # used to count output columns
    for column_index, cell in enumerate(row):
        #if (column_index >= 15 and column_index <= 28): # skip Rain/MSLP
        #    continue
        if (column_index == 1 or column_index == 2): # skip twa of 10 & 20
            offset = offset - 1;
            continue;

        if (cell.isnumeric()): # doesn't cover '.'
            cell = int(cell)

        # we assume wind range is 1 to 40
        if (row_index > firstRow and row_index < lastRow): # the filter data
            #
            if (column_index == 0): # label
                cell = float(cell)
                rowLabel = cell

            if (column_index > 0): # column to process
                col = col + 1
                column_letter = get_column_letter((column_index + 1))

                cell = float(cell)
                rowHrs = rowHrs + cell

                Times[row_index - 9][col - 1] = cell

        if (isinstance(cell, float)): # should add > col 0 so only values
            if (cell > 0.0): # only copy if non-zero & round to 1 decimal place
                # note we can still get a '0.0' as 0.01 is still larger than 0!
                exSheet.cell((row_index + 1), (column_index + offset + 1)).value = cell
                #exSheet.cell((row_index + 1), (column_index + 1)).value = round(cell, 1)
                exSheet.cell((row_index + 1), (column_index + offset + 1)).number_format = '0.0'
                if (column_index > 0):
                    # shade the wind based on hours
                    exSheet.cell((row_index + 1), (column_index + offset + 1)).fill = shadeHrs(cell)
                else:
                    # BFT shade the TWS on the left
                    exSheet.cell((row_index + 1), (column_index + offset + 1)).fill = shadeWind(cell)
        else:
            exSheet.cell((row_index + 1), (column_index + offset + 1)).value = cell
        #exSheet.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

    # now we've processed a row, if a filt row add the total hours to the right
    if (row_index > firstRow and row_index < lastRow): # the filtered data
        if (rowHrs > 0.0): # only add if there is a value
            exSheet.cell((row_index + 1), (column_index + offset + 1 + 1)).value = rowHrs
            exSheet.cell((row_index + 1), (column_index + offset + 1 + 1)).number_format = '0.0'
            exSheet.cell((row_index + 1), (column_index + offset + 1 + 1)).fill = shadeHrs(rowHrs)
        # and add to total
        totHours = totHours + rowHrs

    #print (totWind)
# end of filtered data - conditional shading of green, yellow, red
rule = ColorScaleRule(start_type='percentile', start_value=1, start_color=colors.COLOR_INDEX[3],
                         mid_type='percentile', mid_value=50, mid_color=colors.COLOR_INDEX[5],
                                     end_type='percentile', end_value=99, end_color=colors.COLOR_INDEX[2])
# apply to filtered and totals separately
exSheet.conditional_formatting.add("B10:Q49", rule)
exSheet.conditional_formatting.add("R10:R49", rule)

# add total hours to the lower right
exSheet.cell((lastRow + 1), (lastCol + offset + 1)).value = totHours
exSheet.cell((lastRow + 1), (lastCol + offset + 1)).number_format = '0.0'
# don't shade to avoid confusion
#exSheet.cell((lastRow + 1), (lastCol + 1 + 1)).fill = shadeHrs(totHours)

#
# a lot of the following is in the expedition summary and therefore could be removed
#

# define an array of 6 x 6 for the Summ
Summ = [[0.0] * 6 for i in range(6)]

# total should be 40
Rows = [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, # 22+
        4, 4, 4, 4,  # 18 - 21
        3, 3, 3, 3,  # 14 - 17
        2, 2, 2, 2,  # range 10 - 13
        1, 1, 1, 1, # range 6 - 9
        0, 0, 0, 0, 0, 0] # range 1-5

# and 18 columns
Cols = [0, 0, 0, 0, 0, 0,
        1, 1,
        2, 2,
        3, 3,
        4, 4,
        5, 5, 5, 5]

tot = 0 # do the summing
for row in range (rows):
    for col in range (cols):
        tot = tot + Times[row][col]

        # use a data structure to do the mapping into buckets
        Summ[Rows[row]][Cols[col]] = Summ[Rows[row]][Cols[col]] + Times[row][col]

# location of hr summ 
rowoff = 53
coloff = 5
tabloff = 10

# Add the row col headers for the summary
col = 0
for hdr in ("0-60", "60-80", "80-100", "100-120", "120-140", "140+", "Total"):
    exSheet.cell((rowoff - 1), (coloff + col)).value = hdr
    exSheet.cell((rowoff - 1), (coloff + col)).font = Font(bold = True)
    exSheet.cell((rowoff - 1), (coloff + col)).alignment = Alignment(horizontal = 'right')

    exSheet.cell((rowoff - 1 + tabloff), (coloff + col)).value = hdr
    exSheet.cell((rowoff - 1 + tabloff), (coloff + col)).font = Font(bold = True)
    exSheet.cell((rowoff - 1 + tabloff), (coloff + col)).alignment = Alignment(horizontal = 'right')
    col = col + 1

row = 0
for hdr in ("0 - 6", "6 - 10", "10 - 14", "14 - 18", "18 - 22", "22+", "Total"):
    exSheet.cell((rowoff + row), (coloff - 1)).value = hdr
    exSheet.cell((rowoff + row), (coloff - 1)).font = Font(bold = True)
    exSheet.cell((rowoff + row), (coloff - 1)).alignment = Alignment(horizontal = 'right')

    exSheet.cell((rowoff + row + tabloff), (coloff - 1)).value = hdr
    exSheet.cell((rowoff + row + tabloff), (coloff - 1)).font = Font(bold = True)
    exSheet.cell((rowoff + row + tabloff), (coloff - 1)).alignment = Alignment(horizontal = 'right')
    row = row + 1

# Sum rows - we calc this first so we have it for percentage calc
totHrs = 0.0
for row in range(40):
    #print (row, Times[row][0])
    for col in range (18):
        totHrs = totHrs + Times[row][col]

exSheet.cell((rowoff + 6), (coloff + 6)).value = totHrs
exSheet.cell((rowoff + 6), (coloff + 6)).number_format = '0.00'

# add summ to sheet
for row in range(6):
    tot = 0
    for col in range (6):
        # houra
        exSheet.cell((rowoff + row), (coloff + col)).value = Summ[row][col]
        exSheet.cell((rowoff + row), (coloff + col)).number_format = '0.00'
        # and percentage
        exSheet.cell((rowoff + row + tabloff), (coloff + col)).value = Summ[row][col]/totHrs
        exSheet.cell((rowoff + row + tabloff), (coloff + col)).number_format = '0.0%'
        tot = tot + Summ[row][col]

    exSheet.cell((rowoff + row), (coloff + 6)).value = tot
    exSheet.cell((rowoff + row), (coloff + 6)).number_format = '0.00'
    exSheet.cell((rowoff + row + tabloff), (coloff + 6)).value = tot/totHrs
    exSheet.cell((rowoff + row + tabloff), (coloff + 6)).number_format = '0.0%'

for col in range(6):
    tot = 0
    for row in range (6):
        #exSheet.cell((rowoff + row), (coloff + col)).value = Summ[row][col]
        #exSheet.cell((rowoff + row), (coloff + col)).number_format = '0.00'
        tot = tot + Summ[row][col]
    exSheet.cell((rowoff + 6), (coloff + col)).value = tot
    exSheet.cell((rowoff + 6), (coloff + col)).number_format = '0.00'
    exSheet.cell((rowoff + 6 + tabloff), (coloff + col)).value = tot/totHrs
    exSheet.cell((rowoff + 6 + tabloff), (coloff + col)).number_format = '0.0%'

# Sum rows
totHrs = 0.0
for row in range(40):
    #print (row, Times[row][0])
    for col in range (18):
        totHrs = totHrs + Times[row][col]

exSheet.cell((rowoff + 6), (coloff + 6)).value = totHrs
exSheet.cell((rowoff + 6), (coloff + 6)).number_format = '0.00'

# do conditional formatting for the summaries
exSheet.conditional_formatting.add("E53:J58", rule)
exSheet.conditional_formatting.add("E63:J68", rule)
print (f"Total hours {totHrs}")



# added the avg wind at the bollot
#exSheet.cell((twsCnt + 2), (TWSCOL)).value = 'Average:'
#exSheet.cell((twsCnt + 2), (TWSCOL + 1)).value = twsTotal/twsCnt
#exSheet.cell((twsCnt + 2), (TWSCOL + 1)).number_format = '0.0'
#exSheet.cell((twsCnt + 2), (TWSCOL + 1)).font = Font(bold = True)
# format 1 decimal place - ws['A1'].number_format = '0.00%'

# And the wind buckets - 
clr = 0
#for rng in Angles:
    #per = rng["percent"]/totalDttm
    #exSheet.cell(twsCnt + 6 + clr, 6).value = clr + 1
    #exSheet.cell(twsCnt + 6 + clr, 7).value = str(int(rng["low"])+1) + "-" + str(int(rng["high"]))
    #exSheet.cell(twsCnt + 6 + clr, 8).value = per # rng["percent"]
    #exSheet.cell(twsCnt + 6 + clr, 8).number_format = '00.0%'
    #clr = clr + 1

# added the wind scaling @53 x 12 - the scaling
#exSheet.cell(twsCnt + 6, 11).value = "Wind (Bft)"

#clr = 0
#for clr in range(9):
    #exSheet.cell(twsCnt + 6 + clr, 12).value = clr + 1
    #exSheet.cell(twsCnt + 6 + clr, 13).value = str(int(Winds[clr]["low"])) + "-" + str(int(Winds[clr]["high"])-1)
    #exSheet.cell(twsCnt + 6 + clr, 13).fill = Winds[clr]["fill"]


#print ("TWS ", twsTotal, twsTotal/twsCnt, twsCnt) # Avg wind
# and remove extension
oname, fext = os.path.splitext(fname)
exWorkbook.save(oname + ".xlsx")

# end of file

